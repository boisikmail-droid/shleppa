# -*- coding: utf-8 -*-
"""
Разметка отзывов о заселении/приёмке по смысловым категориям.
Мультиразметка: один отзыв может попасть в несколько категорий.
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SRC = r"C:\vibe\hat1\reviews_utf8.txt"
OUT = r"C:\vibe\hat1\категории_отзывов.xlsx"
OUT_DL = r"c:\Users\viktor\Downloads\Telegram Desktop\категории_отзывов.xlsx"


def load_chunks():
    text = open(SRC, encoding="utf-8").read()
    chunks = [c.strip() for c in re.split(r"---\d+---\n", text)[1:]]
    if chunks and chunks[0].lower().startswith("комментарий"):
        chunks = chunks[1:]
    return chunks


def norm(s: str) -> str:
    s = s.lower().replace("&#34;", '"').replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


def has(t: str, *words: str) -> bool:
    return any(w in t for w in words)


def rx(t: str, pattern: str) -> bool:
    return bool(re.search(pattern, t, flags=re.I))


# id, название, типичный контекст
CATS = [
    (
        "positive",
        "Всё хорошо / положительный без замечаний",
        "всё ок, удобно, быстро, спасибо, понравилось — без жалоб",
    ),
    (
        "app_bad",
        "Приложение/сайт плохо работает (глюки, сбои)",
        "висит, глючит, вылетает, ошибки, не загружается, «приложение говно»",
    ),
    (
        "app_ux",
        "Неудобное приложение / плохой UX",
        "неудобно пользоваться, мелко, неинтуитивно, много окон, с телефона неудобно",
    ),
    (
        "docs_slow",
        "Долгий срок исправления / согласования документов",
        "долго правят документы, недели/месяцы, срыв 72 часов, повторные циклы согласования",
    ),
    (
        "docs_errors",
        "Ошибки в документах / правки не учтены",
        "подсунули старый пакет, Госключ не совпадает с ЛК, правки не внесли",
    ),
    (
        "upsell",
        "Навязывание услуг (оценка, страховка и др.)",
        "нельзя пропустить этап, обязательная оценка/страховка, ненужные сервисы",
    ),
    (
        "keys",
        "Проблемы с ключами / ячейкой / постоматом",
        "ячейка не открылась, неверный номер, долго открывать, постомат не работает",
    ),
    (
        "access",
        "Чип / ключ / домофон / доступ не работает",
        "не открывает подъезд или кладовую, чипы не работают, нет пульта от ворот",
    ),
    (
        "remarks_ux",
        "Неудобно вносить замечания при осмотре",
        "лимит символов, пропала кнопка, нельзя после приёмки, неудобный список дефектов",
    ),
    (
        "navigation",
        "Навигация на объекте (кладовая, -1, паркинг)",
        "непонятно куда идти, нет указателей, лифт на -1, кладовая в другом подъезде",
    ),
    (
        "wifi",
        "Wi-Fi / интернет на объекте мешает процессу",
        "требуют Wi-Fi ПИК, связь глушат, без интернета нельзя открыть ячейку",
    ),
    (
        "schedule",
        "Долгая запись / ожидание / срыв сроков выдачи",
        "далеко ближайшая дата осмотра, ждали менеджера, ключи месяцами",
    ),
    (
        "defects",
        "Дефекты квартиры / отделка / инженерия",
        "качество отделки, замок, счётчики, нет горячей воды, грязь в квартире",
    ),
    (
        "support",
        "Менеджер / поддержка / связь",
        "не позвонили, чат бесполезен, ждали сотрудника, нет ответа",
    ),
    (
        "unclear",
        "Непонятный процесс / нет инструкций",
        "непрозрачно, нет информации, без консультации не разобраться",
    ),
    (
        "process_bad",
        "Плохая организация процесса заселения",
        "ужасная процедура, этапы плохо организованы, пользовательский опыт отвратительный",
    ),
    (
        "dirt",
        "Грязь / мусор в МОП / на объекте",
        "грязно в подъезде, строительный мусор",
    ),
    (
        "no_staff",
        "Нет представителя на месте",
        "один на один, некому показать объект, отсутствие представителя",
    ),
    (
        "other_neg",
        "Прочий негатив без деталей",
        "общий негатив без конкретики («ужас», «отвратительно»)",
    ),
    (
        "neutral",
        "Нейтрально / неинформативно",
        "пусто, точка, одно слово без смысла",
    ),
]

NAME = {c[0]: c[1] for c in CATS}
CTX = {c[0]: c[2] for c in CATS}

POS_EXACT = {
    "все ок", "всё ок", "все хорошо", "всё хорошо", "все отлично", "всё отлично",
    "отлично", "супер", "удобно", "очень удобно", "быстро и удобно", "быстро и понятно",
    "все удобно", "всё удобно", "все понятно", "всё понятно", "все понравилось",
    "всё понравилось", "спасибо", "благодарю", "ок", "нормально", "всё нормально",
    "все нормально", "ничего", "все", "всё", "пойдет", "пойдёт", "хорошо", "удобство",
    "скорость", "онлайн", "все гуд", "всё гуд", "всё супер", "все супер",
    "идеально удобно", "без проблем", "норм", "гуд", "good", "ok", "быстро", "понятно",
    "удобно быстро понятно", "удобно быстро", "быстро удобно", "чисто и уютно",
    "бесконтактное заселение никто не мешал при осмотре",
}

NEG_HARD = [
    "ужас", "отврат", "говн", "кошмар", "безобразие", "криво", "плохо",
    "не работ", "завис", "зависа", "глюч", "навяз", "неудоб", "не удоб",
    "ошибк", "сбо", "лага", "висит", "виснет", "вылет", "тормоз", "ад!!!",
]

TECH_BAD = [
    "завис", "зависа", "висит", "виснет", "глюч", "вылет", "выкид", "выбрас",
    "сбо", "ошибк", "лага", "тормоз", "не загруж", "нестабиль", "не груз",
    "не прогруз", "сырой", "техническ", "обновляться несколько",
]

PLATFORM = [
    "приложен", "сайт", "лк", "личн", "кабинет", "веб", "онлайн", "сервис",
]


def is_positive(t: str) -> bool:
    clean = t.strip(" !.").strip()
    if t in POS_EXACT or clean in POS_EXACT:
        return True
    if len(t) <= 100 and has(
        t, "удобно", "отлично", "супер", "хорошо", "спасибо", "понравил",
        "понятно", "быстро", "благодар", "идеальн", "молодц", "гуд", "норм",
        "оператив", "в пользу собственник",
    ):
        if has(t, "неудоб", "не удоб", "завис", "зависа", "глюч", "ужас", "отврат",
               "плох", "навяз", "говн", "висит", "виснет", "ошибк", "сбо"):
            return False
        if has(t, "но ", "кроме", "только лучше", "только жаль") and has(t, *NEG_HARD):
            return False
        # «удобно. правки согласовали» — позитив
        return True
    return False


def is_app_bad(t: str) -> bool:
    tech = has(t, *TECH_BAD)
    platform = has(t, *PLATFORM)
    if platform and tech:
        return True
    if platform and has(t, "говн", "ужас", "отврат", "крив", "кошмар", "не работ", "плох"):
        return True
    # короткие жалобы на зависание без слова «приложение»
    if tech and len(t) <= 120 and has(t, "висит", "виснет", "завис", "зависа", "глюч", "сбо"):
        return True
    if has(t, "выкидывает из документ", "выбрасывает из документ", "прочитать документы практически невозможно"):
        return True
    if has(t, "корректную ссылку", "ссылку для начала заселения") and has(t, "недел", "долг", "ошиб"):
        return True
    return False


def is_app_ux(t: str) -> bool:
    if has(
        t,
        "неудобн", "не удобн", "не интуитив", "нелогичн", "мелко", "масштаб",
        "не адаптир", "с телефона", "много окон", "всплывающ",
        "абсолютно неудоб", "крайне не удоб", "вообще неудоб", "очень не удоб",
        "очень неудоб", "электронный формат мне не удоб", "не комфортн",
        "не клиенто", "пользовательский опыт",
    ):
        return True
    if has(t, "интерфейс") and has(t, *NEG_HARD + ["плох", "доработ", "неудоб"]):
        return True
    return False


def is_docs_slow(t: str) -> bool:
    if rx(t, r"(долг|недел|месяц|час|ждал|ждали|72).{0,60}(документ|соглас|правк|пакет|исправл)"):
        return True
    if rx(t, r"(документ|соглас|правк|пакет|исправл).{0,60}(долг|недел|месяц|час|ждал|ждали|72)"):
        return True
    if has(t, "более 72", "почти 2 недель", "почти две недель", "2 недель", "две недель", "3 недель"):
        return True
    if has(t, "повтор", "снова", "третий раз", "пятый раз", "который раз", "дважды") and has(
        t, "документ", "соглас", "правк", "пакет", "направляю",
    ):
        return True
    if has(t, "правили", "вносили изменения") and has(t, "долг", "недел", "месяц"):
        return True
    return False


def is_docs_errors(t: str) -> bool:
    if not has(t, "документ", "паспорт", "госключ", "пакет", "правк", "согласован", "эцп", "скан", "гос. ключ", "госключ"):
        return False
    if is_docs_slow(t):
        return False
    if is_positive(t) and len(t) < 100 and not has(t, "ошиб", "не ", "правк"):
        return False
    return has(
        t,
        "ошиб", "правк", "повтор", "вернул", "подсун", "подсов", "не учл",
        "не исправ", "снова", "госключ", "гос. ключ", "без учета", "без учёта",
        "пришлось", "некоррект", "не отображ", "не прислал", "список документов",
        "скачать все документы", "архивом", "замечания не были учтены",
        "корректировки не вносятся",
    )


def is_upsell(t: str) -> bool:
    if has(t, "навяз"):
        return True
    if has(t, "ненужн", "не нужн", "лишн") and has(t, "сервис", "услуг", "этап", "пункт"):
        return True
    if has(t, "оценк", "страхов") and has(
        t, "нельзя", "обязательн", "принужд", "пропуст", "без возможности",
        "кнопки пропустить", "навяз",
    ):
        return True
    if has(t, "без возможности отказаться"):
        return True
    return False


def is_keys(t: str) -> bool:
    if has(t, "ячейк", "ящик", "постомат", "постамат"):
        return True
    if has(t, "ключ") and has(
        t, "получ", "выдач", "открыт", "не откры", "ячей", "ящик", "не работа",
        "ждал", "ждали", "неверн", "другом подъезд", "не можем принять",
    ):
        return True
    return False


def is_support(t: str) -> bool:
    if has(t, "не позвон", "менеджер не", "сотрудник не", "искали менеджер", "ждали менеджер"):
        return True
    if has(t, "менеджер", "поддерж", "чат", "оператор", "горяч", "контакт центр", "обратной связи"):
        if has(t, *NEG_HARD + [
            "не позвон", "не звон", "ждал", "ждали", "не помог", "бесполез",
            "отписк", "нет ответа", "нет конкретн", "дублировать", "вода",
        ]):
            return True
    return False


def is_process_bad(t: str) -> bool:
    return has(
        t,
        "ужасная процедура", "ужасный процесс", "ужасный сервис", "отвратительный сервис",
        "плохо организова", "отвратительно организу", "не продуман", "бизнес процесс",
        "криво и тупо", "7 кругов", "электронный формат", "все этапы",
        "организац",
    ) and has(t, *NEG_HARD + ["плох", "ужас", "отврат", "крив", "неудоб", "замороч"])


def is_unclear(t: str) -> bool:
    return has(
        t,
        "непонят", "не понят", "нет инструкц", "нет информац", "нет полной",
        "непрозрач", "не хватает", "недостаточно", "без консультац", "неясн",
        "где список", "сначала нужно осмотреть",
    )


def is_no_staff(t: str) -> bool:
    return has(
        t,
        "нет сотрудник", "нет представител", "один на один", "некому",
        "никто не объясн", "отсутствие представител", "не нравится отсутствие представител",
    )


def label(raw: str) -> list[str]:
    t = norm(raw)
    labels: list[str] = []

    if is_positive(t):
        labels.append("positive")
    if is_app_bad(t):
        labels.append("app_bad")
    if is_app_ux(t):
        labels.append("app_ux")
    if is_docs_slow(t):
        labels.append("docs_slow")
    if is_docs_errors(t):
        labels.append("docs_errors")
    if is_upsell(t):
        labels.append("upsell")
    if is_keys(t):
        labels.append("keys")
    if has(t, "чип", "домофон", "пульт", "ворот") or has(t, "ключ не работа", "не открывает подъезд"):
        labels.append("access")
    if has(
        t, "символ", "лимит", "150 знак", "внести замечан", "пропала кнопка",
        "замечания по одном", "после приемки", "после приёмки",
    ) or (has(t, "замечан") and has(t, "неудоб", "не удоб", "нельзя", "пропал", "символ", "лимит", "кнопк")):
        labels.append("remarks_ux")
    if has(
        t, "кладов", "-1", "−1", "парков", "машиномес", "навигац", "указател",
        "карта не", "второй подъезд", "куда идти", "как пройти", "как попасть", "метку",
    ):
        labels.append("navigation")
    if has(t, "вайф", "wi-fi", "wifi", "wi‑fi", "глуш"):
        labels.append("wifi")
    if has(
        t, "запис", "ближайш", "свободных дат", "интервал на запись",
        "срываете срок", "2,5 месяц", "быстрее прийти на осмотр",
    ) or (has(t, "ждал", "ждали", "искали") and has(t, "час", "менеджер")):
        labels.append("schedule")
    if has(
        t, "дефект", "отделк", "качество отдел", "дверной замок", "плесень",
        "горячей воды", "матрас", "счётчик", "счетчик",
    ):
        labels.append("defects")
    if is_support(t):
        labels.append("support")
    if is_unclear(t):
        labels.append("unclear")
    if is_process_bad(t):
        labels.append("process_bad")
    if has(t, "грязн", "мусор", "строительн мусор"):
        labels.append("dirt")
    if is_no_staff(t):
        labels.append("no_staff")
    if has(t, "лифт не работа"):
        labels.append("navigation")

    # чистый позитив
    if "positive" in labels and not any(
        x in labels
        for x in [
            "app_bad", "app_ux", "docs_slow", "docs_errors", "upsell", "keys",
            "access", "remarks_ux", "navigation", "wifi", "schedule", "defects",
            "support", "unclear", "process_bad", "dirt", "no_staff",
        ]
    ):
        return ["positive"]

    # убрать positive если есть сильный негатив
    if "positive" in labels and any(
        x in labels for x in ["app_bad", "docs_slow", "upsell", "keys", "process_bad"]
    ):
        labels = [x for x in labels if x != "positive"]

    if not labels:
        if len(t) < 3 or t in {".", "!", "-", "...", "?"}:
            return ["neutral"]
        if has(t, *NEG_HARD):
            return ["other_neg"]
        if len(t) < 45:
            return ["neutral"]
        return ["other_neg"]

    return labels


PRIORITY = [
    "app_bad", "app_ux", "docs_slow", "docs_errors", "upsell", "keys", "access",
    "remarks_ux", "navigation", "wifi", "schedule", "defects", "support",
    "unclear", "process_bad", "dirt", "no_staff", "other_neg", "positive", "neutral",
]


def build_workbook(chunks):
    multi: dict[str, list[str]] = defaultdict(list)
    primary: dict[str, list[str]] = defaultdict(list)
    detail = []

    for i, raw in enumerate(chunks, 1):
        labs = label(raw)
        prim = next(p for p in PRIORITY if p in labs)
        primary[prim].append(raw)
        for L in set(labs):
            multi[L].append(raw)
        detail.append((i, prim, labs, raw))

    total = len(chunks)

    # umbrella sets
    app_all = set(multi["app_bad"]) | set(multi["app_ux"])
    docs_all = set(multi["docs_slow"]) | set(multi["docs_errors"])

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    accent_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # --- Сводка ---
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Категория", "Типичный контекст отзыва", "Кол-во", "% от всех"])
    style_header(ws, 4)

    # Сводные (как в примере пользователя) — сверху
    umbrella = [
        (
            "Всё хорошо (смысл «всё ок / удобно / спасибо»)",
            "короткий позитив без жалоб",
            len(multi["positive"]),
        ),
        (
            "Приложение плохое / не работает (глюки + «говно»)",
            "технические сбои приложения/сайта/ЛК",
            len(multi["app_bad"]),
        ),
        (
            "Неудобное приложение (UX)",
            "неудобно, неинтуитивно, мелко, плохой интерфейс",
            len(multi["app_ux"]),
        ),
        (
            "Проблемы с приложением/сайтом ВСЕГО (глюки + UX)",
            "объединение: плохо работает ИЛИ неудобно",
            len(app_all),
        ),
        (
            "Долгий срок исправления документов",
            "долго правят / повторные циклы согласования / срыв сроков",
            len(multi["docs_slow"]),
        ),
        (
            "Документы: ошибки + долго (всё про документы)",
            "объединение docs_slow + docs_errors",
            len(docs_all),
        ),
    ]

    for name, ctx, n in umbrella:
        row = [name, ctx, n, round(100.0 * n / total, 1)]
        ws.append(row)
        for col in range(1, 5):
            ws.cell(ws.max_row, col).fill = accent_fill

    ws.append(["—", "— детализация по всем категориям ниже —", "", ""])

    multi_rows = sorted(
        ((cid, NAME[cid], CTX[cid], len(multi[cid])) for cid, _, _ in CATS),
        key=lambda x: -x[3],
    )
    for cid, name, ctx, n in multi_rows:
        if n == 0:
            continue
        ws.append([name, ctx, n, round(100.0 * n / total, 1)])

    ws.append([])
    ws.append([
        "Всего уникальных отзывов",
        "Один отзыв может быть в нескольких категориях — сумма строк > итога",
        total,
        100,
    ])
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = thin

    # --- Без пересечений ---
    ws2 = wb.create_sheet("Без пересечений")
    ws2.append(["Категория", "Типичный контекст отзыва", "Кол-во", "%"])
    style_header(ws2, 4)
    for cid, name, ctx, _ in multi_rows:
        n = len(primary[cid])
        if n:
            ws2.append([name, ctx, n, round(100.0 * n / total, 1)])
    ws2.append([])
    ws2.append(["Итого", "Каждый отзыв учтён один раз", total, 100])
    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 70
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 10

    # --- Примеры ---
    ws3 = wb.create_sheet("Примеры")
    ws3.append(["Категория", "Кол-во", "Пример отзыва"])
    style_header(ws3, 3)
    for cid, name, ctx, n in multi_rows:
        if not n:
            continue
        for ex in multi[cid][:10]:
            ws3.append([name, n, ex[:700]])
    ws3.column_dimensions["A"].width = 48
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 100

    # --- Все ---
    ws4 = wb.create_sheet("Все отзывы")
    ws4.append(["№", "Основная категория", "Все категории", "Текст"])
    style_header(ws4, 4)
    for i, prim, labs, raw in detail:
        ws4.append([i, NAME[prim], "; ".join(NAME[x] for x in labs), raw])
    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["C"].width = 55
    ws4.column_dimensions["D"].width = 100

    wb.save(OUT)
    try:
        wb.save(OUT_DL)
    except Exception as e:
        print("DL save skip:", e)

    return total, multi_rows, umbrella, multi


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    chunks = load_chunks()
    total, multi_rows, umbrella, multi = build_workbook(chunks)

    print(f"Всего отзывов: {total}\n")
    print("=== Сводка (как в примере) ===")
    for name, ctx, n in umbrella:
        print(f"{n:4d}  {name}")
        print(f"      {ctx}")
    print("\n=== Все категории ===")
    for cid, name, ctx, n in multi_rows:
        if n:
            print(f"{n:4d}  {name}")
    print(f"\nФайл: {OUT}")
    print(f"Копия: {OUT_DL}")


if __name__ == "__main__":
    main()
